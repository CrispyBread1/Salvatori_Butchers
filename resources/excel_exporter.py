from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Border, Side
from PyQt5.QtWidgets import QFileDialog
from openpyxl.worksheet.page import PageMargins

class ExcelExporter:
    def __init__(self, parent=None):
        self.parent = parent

    def export(
        self,
        data,
        group_by=None,
        sheet_name="Sheet1",
        title=None,
        filename=None,
        headers=None,
        butchers_list=None
    ):
        """
        Export list-of-dict data to Excel.

        :param data: List[Dict], input data
        :param group_by: str | None, key to group by (e.g., 'customer_name')
        :param sheet_name: str, name of the Excel sheet
        :param filename: str | None, if None shows save dialog
        :param headers: List[str] | None, if None uses keys from first dict
        """
        if not data or not isinstance(data, list):
            raise ValueError("Data must be a non-empty list of dictionaries.")

        

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        row = 1
        title_font = Font(
            bold=True,
            size=20
        )
        bold_font = Font(
            bold=True,
            size=12
        )
        border = Border(
            left=Side(border_style="thin",
                color='00000000'),
            right=Side(border_style="thin",
                color='00000000'),
            top=Side(border_style="thin",
                color='00000000'),
            bottom=Side(border_style="thin",
                color='00000000')
        )



        if title:
            ws.cell(row=1, column=1, value=title).font = title_font
            row += 1

        if group_by:
            grouped = {}
            keys = headers or list(items[0].keys())

            for col, key in enumerate(keys, start=1):
                ws.cell(row=row, column=col, value=key).font = bold_font
            if butchers_list:
                ws.cell(row=row, column=(len(keys) + 1), value="Weight").font = bold_font
                ws.cell(row=row, column=(len(keys) + 2), value="Code").font = bold_font
            row += 1

            for item in data:
                key = item.get(group_by, "Ungrouped")
                grouped.setdefault(key, []).append(item)



            for group, items in grouped.items():
                ws.cell(row=row, column=1, value=f"{group_by.capitalize()}: {group}").font = bold_font
                row += 1


                for item in items:
                    for col, key in enumerate(keys, start=1):
                        ws.cell(row=row, column=col, value=item.get(key.lower(), ""))
                    if butchers_list:
                        ws.cell(row=row, column=(len(keys) + 1), value="").border = border
                        ws.cell(row=row, column=(len(keys) + 2), value="").border = border
                    row += 1
                
                ws.row_dimensions[row].height = 4
                row += 1  # spacer between groups
                
        else:
            keys = headers or list(data[0].keys())
            for col, key in enumerate(keys, start=1):
                ws.cell(row=row, column=col, value=key).font = bold_font
            row += 1

            for item in data:
                for col, key in enumerate(keys, start=1):
                    ws.cell(row=row, column=col, value=item.get(key, ""))
                row += 1

        if not filename:
            filename, _ = QFileDialog.getSaveFileName(
                self.parent, "Save Excel File", "", "Excel Files (*.xlsx)"
            )
            if not filename:
                return
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"


        if butchers_list:
            for column_cells in ws.columns:
              length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
              ws.column_dimensions[column_cells[0].column_letter].width = length

            ws.page_margins = PageMargins(
                left=0.23622,
                right=0.23622,
                top=0.15748,
                bottom=0.15748,
                header=0.3,
                footer=0.3
            )


        wb.save(filename)
